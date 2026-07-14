#!/usr/bin/env python3
"""
step0_dicom2nii.py

Converts XNAT-downloaded DICOM data to BIDS-formatted NIfTI files in sourcedata/.

Stage 0  — Reads the subject Excel tracking file + XNAT scan folder structure,
           then writes scan_mapping.json to sourcedata/sub-XX/ses-YY/.
           If the JSON already exists Stage 0 is skipped (--remap forces regen).

Stage 1  — Reads scan_mapping.json, merges magnitude + phase DICOMs for each
           run into a temporary directory, calls dcm2niix -b y -z y, renames
           outputs to BIDS sourcedata convention, and injects TaskName into
           functional JSON sidecars.

Usage:
    python step0_dicom2nii.py --sub 01 --ses 01
    python step0_dicom2nii.py --sub 01 --ses 01 --xnat-dir /path/to/HD01_S01_070726
    python step0_dicom2nii.py --sub 01 --ses 01 --remap          # regenerate JSON
    python step0_dicom2nii.py --sub 01 --ses 01 --stage0-only    # JSON only, no conversion
"""

import argparse
import glob
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile

import openpyxl

# ---------------------------------------------------------------------------
# Hard-coded project paths  (edit these for your environment)
# ---------------------------------------------------------------------------
XNAT_BASE_DIR  = '/Users/anatriana/Documents/7T/'
SOURCEDATA_DIR = '/Users/anatriana/Documents/7T/sourcedata'   # TODO: update to real path
# Excel filename uses un-padded subject number: sub-1.xlsx, sub-2.xlsx, …
EXCEL_TEMPLATE = os.path.join(XNAT_BASE_DIR, 'sub-{sub_int}.xlsx')

# ---------------------------------------------------------------------------
# Task label mapping: scanner / Excel column-B label  →  BIDS task name
# Labels that map to None are ambiguous and must be resolved in the Excel
# (experimenter must write the full disambiguated name in column B).
# ---------------------------------------------------------------------------
TASK_MAP = {
    # Short scanner labels
    'ava':       'audvisattn',
    'avwm':      'audviswm',
    'proj':      'epiproj',
    'mdspatial': 'spatialwm',
    'mdverbal':  'verbalwm',
    'mdvmsit':   'vmsit',
    'langloc':   None,   # ambiguous — must be langlocaud or langlocvis in Excel col B
    'tom':       None,   # ambiguous — must be tomfalse or tompain in Excel col B
    # Full BIDS names (already correct — pass through)
    'msit':       'msit',
    'rest':       'rest',
    'audvisattn': 'audvisattn',
    'audviswm':   'audviswm',
    'epiproj':    'epiproj',
    'spatialwm':  'spatialwm',
    'verbalwm':   'verbalwm',
    'vmsit':      'vmsit',
    'langlocaud': 'langlocaud',
    'langlocvis': 'langlocvis',
    'tomfalse':   'tomfalse',
    'tompain':    'tompain',
}

# For each BIDS task name, the set of possible scanner labels that may appear
# in XNAT folder names (handles old short labels and new full BIDS labels).
BIDS_TO_SCANNER_LABELS = {
    'audvisattn': {'ava',       'audvisattn'},
    'audviswm':   {'avwm',      'audviswm'},
    'epiproj':    {'proj',      'epiproj'},
    'spatialwm':  {'mdspatial', 'spatialwm'},
    'verbalwm':   {'mdverbal',  'verbalwm'},
    'vmsit':      {'mdvmsit',   'vmsit'},
    'langlocaud': {'langloc',   'langlocaud'},
    'langlocvis': {'langloc',   'langlocvis'},
    'tomfalse':   {'tom',       'tomfalse'},
    'tompain':    {'tom',       'tompain'},
    'msit':       {'msit'},
    'rest':       {'rest'},
}

# XNAT folder-name keywords that identify anatomical series → BIDS suffix
ANAT_KEYWORDS = {
    'INV1':       'inv-1_MP2RAGE',
    'INV2':       'inv-2_MP2RAGE',
    'UNI_Images': 'UNIT1',
}

# Excel column-B labels that denote anatomical (non-functional) acquisitions
ANAT_TASK_LABELS = {'t1', 't2', 'qsm'}

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description='Convert XNAT DICOMs to BIDS sourcedata (step 0).')
    p.add_argument('--sub',  required=True,
                   help='Subject number, zero-padded (e.g. 01)')
    p.add_argument('--ses',  required=True,
                   help='Session number, zero-padded (e.g. 01)')
    p.add_argument('--xnat-dir', default=None,
                   help='Explicit path to XNAT session folder (skips auto-discovery)')
    p.add_argument('--remap', action='store_true',
                   help='Force regeneration of scan_mapping.json even if it exists')
    p.add_argument('--stage0-only', action='store_true',
                   help='Build scan_mapping.json only; skip DICOM conversion')
    return p.parse_args()

# ---------------------------------------------------------------------------
# XNAT session folder discovery
# ---------------------------------------------------------------------------

def find_xnat_dir(sub, ses):
    """
    Search XNAT_BASE_DIR for a session folder matching subject and session numbers.
    Handles naming variants:
        HD01_01_DDMMYY   (original format)
        HD1_S02          (one-off mistake, no leading zero, no date)
        HD01_S03_DDMMYY  (new standard format)
    Raises if 0 or >1 folders match.
    """
    sub_int = int(sub)
    ses_int = int(ses)

    # Patterns: subject number with optional leading zero, session with optional S prefix
    patterns = [
        rf'^HD0?{sub_int}_S0?{ses_int}(_\d+)?$',   # HD01_S01 or HD01_S01_DDMMYY
        rf'^HD0?{sub_int}_0?{ses_int}_\d+$',         # HD01_01_DDMMYY (old format)
    ]

    try:
        entries = [e for e in os.listdir(XNAT_BASE_DIR)
                   if os.path.isdir(os.path.join(XNAT_BASE_DIR, e))]
    except FileNotFoundError:
        sys.exit(f'ERROR: XNAT_BASE_DIR not found: {XNAT_BASE_DIR}')

    matches = []
    for entry in entries:
        for pat in patterns:
            if re.match(pat, entry, re.IGNORECASE):
                matches.append(entry)
                break

    if len(matches) == 0:
        sys.exit(
            f'ERROR: No XNAT session folder found for sub-{sub} ses-{ses} '
            f'in {XNAT_BASE_DIR}.\n'
            f'Use --xnat-dir to specify the folder explicitly.'
        )
    if len(matches) > 1:
        sys.exit(
            f'ERROR: Multiple XNAT session folders match sub-{sub} ses-{ses}:\n'
            f'  {matches}\n'
            f'Rename the duplicates or use --xnat-dir.'
        )

    return os.path.join(XNAT_BASE_DIR, matches[0])

# ---------------------------------------------------------------------------
# XNAT scan inventory
# ---------------------------------------------------------------------------

def inventory_xnat_scans(xnat_dir):
    """
    Walk xnat_dir/scans/ and return a list of dicts (sorted by scan number)
    describing every scan subfolder.
    """
    scans_root = os.path.join(xnat_dir, 'scans')
    if not os.path.isdir(scans_root):
        sys.exit(f'ERROR: No scans/ directory under {xnat_dir}')

    scans = []
    for entry in os.listdir(scans_root):
        if not os.path.isdir(os.path.join(scans_root, entry)):
            continue
        m = re.match(r'^(\d+)-(.+)$', entry)
        if not m:
            continue

        scan_num  = int(m.group(1))
        scan_name = m.group(2)

        is_physio   = 'PhysioLog' in scan_name
        is_fieldmap = 'invPE' in scan_name
        is_anat     = any(k in scan_name for k in ANAT_KEYWORDS)
        is_localizer = re.search(r'localiz', scan_name, re.IGNORECASE) is not None

        # Extract the task token (first underscore-delimited component, lowercased)
        task_in_folder = None
        base_name = scan_name.replace('_PhysioLog', '')
        if not (is_fieldmap or is_anat or is_localizer):
            task_in_folder = base_name.split('_')[0].lower()

        scans.append({
            'scan_num':       scan_num,
            'folder':         entry,
            'scan_name':      scan_name,
            'is_physio':      is_physio,
            'is_fieldmap':    is_fieldmap,
            'is_anat':        is_anat,
            'is_localizer':   is_localizer,
            'task_in_folder': task_in_folder,
            'dicom_dir':      os.path.join(
                scans_root, entry, 'resources', 'DICOM', 'files'),
        })

    return sorted(scans, key=lambda x: x['scan_num'])

# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_excel(excel_path, ses):
    """
    Parse the ses-N sheet from the subject tracking Excel.
    Returns a list of dicts with keys:
        task_label  — raw label from column B (lowercased)
        bids_task   — resolved BIDS task name
        run_num     — integer run number from column C
        successful  — True if col D == 'Y' and col F == 'Y'
        scan_name   — descriptive name from column A
    Only functional-task rows are returned (anatomicals and None-label rows skipped).
    """
    if not os.path.isfile(excel_path):
        sys.exit(f'ERROR: Excel file not found: {excel_path}')

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = f'ses-{int(ses)}'
    if sheet_name not in wb.sheetnames:
        sys.exit(f'ERROR: Sheet "{sheet_name}" not found in {excel_path}')

    ws = wb[sheet_name]

    # Locate the header row: the row where column B (index 1) == 'task'
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[1] is not None and str(row[1]).strip().lower() == 'task':
            header_row_idx = i + 1  # convert to 1-based for min_row
            break

    if header_row_idx is None:
        sys.exit(f'ERROR: Cannot find header row with "task" in column B '
                 f'in sheet {sheet_name}')

    runs = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(v is not None for v in row):
            continue  # blank row

        scan_name  = row[0]
        task_label = row[1]
        run_num    = row[2]
        scan_ok    = row[3]   # col D: scan successful (technical)?
        task_ok    = row[5]   # col F: task successful (technical)?

        if task_label is None:
            continue  # Localizer, InvPE fieldmap rows have no task label

        task_label = str(task_label).strip().lower()

        if task_label in ANAT_TASK_LABELS:
            continue  # anatomicals are discovered from XNAT folder names

        # Resolve BIDS task name
        if task_label not in TASK_MAP:
            print(f'  WARNING: unrecognised task label "{task_label}" in '
                  f'{sheet_name} — using as-is')
            bids_task = task_label
        else:
            bids_task = TASK_MAP[task_label]
            if bids_task is None:
                # Ambiguous label that was never resolved in the Excel
                sys.exit(
                    f'ERROR: Ambiguous task label "{task_label}" in {sheet_name}.\n'
                    f'Update column B to the specific BIDS name '
                    f'(tomfalse/tompain or langlocaud/langlocvis).'
                )

        successful = (
            str(scan_ok).strip().upper() == 'Y' and
            str(task_ok).strip().upper() == 'Y'
        )

        runs.append({
            'task_label': task_label,
            'bids_task':  bids_task,
            'run_num':    int(run_num) if run_num is not None else None,
            'successful': successful,
            'scan_name':  str(scan_name) if scan_name else '',
        })

    return runs

# ---------------------------------------------------------------------------
# Stage 0: build scan_mapping.json
# ---------------------------------------------------------------------------

def build_scan_mapping(sub, ses, xnat_dir, excel_runs, xnat_scans):
    """
    Match Excel rows to XNAT scan folders and return the mapping dict.
    The dict is serialised to scan_mapping.json and consumed by Stage 1.
    """
    mapping = {
        'sub':              sub,
        'ses':              ses,
        'xnat_session_dir': os.path.basename(xnat_dir),
        'generated_by':     'step0_dicom2nii.py',
        'func':             [],
        'fmap':             [],
        'anat':             [],
    }

    # ── Fieldmaps ──────────────────────────────────────────────────────────
    fmap_func   = [s for s in xnat_scans if s['is_fieldmap'] and not s['is_physio']]
    fmap_physio = [s for s in xnat_scans if s['is_fieldmap'] and s['is_physio']]

    if fmap_func:
        fmap_entry = {
            'xnat_scan_dirs': [s['folder'] for s in fmap_func],
            'dir': 'PA',
        }
        if fmap_physio:
            fmap_entry['physio_dir'] = fmap_physio[0]['folder']
        mapping['fmap'].append(fmap_entry)

    # ── Anatomicals ─────────────────────────────────────────────────────────
    for scan in xnat_scans:
        if not scan['is_anat']:
            continue
        for keyword, bids_suffix in ANAT_KEYWORDS.items():
            if keyword in scan['scan_name']:
                mapping['anat'].append({
                    'xnat_scan_dir': scan['folder'],
                    'bids_suffix':   bids_suffix,
                })
                break

    # ── Functional runs ─────────────────────────────────────────────────────
    func_scans  = [s for s in xnat_scans
                   if not s['is_physio'] and not s['is_fieldmap']
                   and not s['is_anat'] and not s['is_localizer']
                   and s['task_in_folder'] is not None]

    func_physio = [s for s in xnat_scans
                   if s['is_physio'] and not s['is_fieldmap']
                   and s['task_in_folder'] is not None]

    # Iterate task labels in the order they first appear in the Excel
    seen_tasks = []
    for r in excel_runs:
        if r['task_label'] not in seen_tasks:
            seen_tasks.append(r['task_label'])

    for task_label in seen_tasks:
        excel_rows = [r for r in excel_runs if r['task_label'] == task_label]
        bids_task  = excel_rows[0]['bids_task']

        # XNAT folders whose task token matches any known scanner label for this task
        possible = BIDS_TO_SCANNER_LABELS.get(bids_task, {bids_task, task_label})
        task_xnat   = [s for s in func_scans  if s['task_in_folder'] in possible]
        task_physio = [s for s in func_physio if s['task_in_folder'] in possible]

        n_excel = len(excel_rows)
        n_xnat  = len(task_xnat)

        # Group XNAT functional folders into per-run groups.
        # Expected: 2 folders per run (magnitude + phase).
        if n_xnat == n_excel * 2:
            groups = [task_xnat[i * 2: i * 2 + 2] for i in range(n_excel)]
        elif n_xnat == n_excel:
            # Unusual: only 1 series per run (e.g. failed run cut short, mock data)
            print(f'  WARNING: {n_xnat} XNAT folder(s) for "{task_label}" but '
                  f'expected {n_excel * 2} (2 per run). Assigning 1 per run.')
            groups = [[task_xnat[i]] for i in range(n_excel)]
        else:
            # Mismatch: assign greedily and warn
            print(f'  WARNING: cannot cleanly pair {n_xnat} XNAT folders to '
                  f'{n_excel} Excel rows for "{task_label}". Assigning greedily.')
            groups, idx = [], 0
            for _ in range(n_excel):
                chunk_size = min(2, n_xnat - idx)
                groups.append(task_xnat[idx: idx + chunk_size])
                idx += chunk_size

        for excel_row, xnat_group in zip(excel_rows, groups):
            run_str = f'{excel_row["run_num"]:02d}'
            entry = {
                'xnat_scan_dirs': [s['folder'] for s in xnat_group],
                'bids_task':      bids_task,
                'run':            run_str,
                'include':        excel_row['successful'],
            }

            if not excel_row['successful']:
                entry['reason'] = 'scan_or_task_failed'

            # Associate the nearest physio folder by scan number
            if task_physio:
                max_func_num = max(s['scan_num'] for s in xnat_group)
                nearest = min(task_physio,
                              key=lambda p: abs(p['scan_num'] - max_func_num))
                entry['physio_dir'] = nearest['folder']

            mapping['func'].append(entry)

    return mapping

# ---------------------------------------------------------------------------
# Stage 1 helpers
# ---------------------------------------------------------------------------

def _dicom_dir(xnat_dir, scan_folder):
    return os.path.join(xnat_dir, 'scans', scan_folder, 'resources', 'DICOM', 'files')


def merge_dicoms(xnat_dir, scan_folders, dest_dir):
    """
    Copy DICOMs from one or more XNAT scan folders into a single directory so
    dcm2niix sees both magnitude and phase and can split them via ImageType.
    """
    for folder in scan_folders:
        src = _dicom_dir(xnat_dir, folder)
        if not os.path.isdir(src):
            print(f'    WARNING: DICOM directory not found: {src}')
            continue
        for fname in os.listdir(src):
            fpath = os.path.join(src, fname)
            if os.path.isfile(fpath):
                shutil.copy2(fpath, os.path.join(dest_dir, fname))


def run_dcm2niix(dicom_dir, out_dir):
    """Run dcm2niix with BIDS sidecar and gzip on dicom_dir → out_dir."""
    cmd = ['dcm2niix', '-b', 'y', '-z', 'y', '-f', 'dcm2niix_out', '-o', out_dir, dicom_dir]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f'dcm2niix failed:\n{result.stderr}')


def is_phase_file(fname):
    """
    dcm2niix appends _ph (and possible extra suffixes) to phase images.
    Detect by checking whether '_ph' appears in the stem (before the extension).
    """
    stem = fname
    for ext in ('.nii.gz', '.nii', '.json'):
        if stem.endswith(ext):
            stem = stem[: -len(ext)]
    # _ph may appear at end of stem or followed by another underscore
    return bool(re.search(r'_ph(_|$)', stem, re.IGNORECASE))


def inject_task_name(json_path, task_name):
    if not os.path.isfile(json_path):
        return
    with open(json_path) as f:
        data = json.load(f)
    data['TaskName'] = task_name
    with open(json_path, 'w') as f:
        json.dump(data, f, indent=2)


# ---------------------------------------------------------------------------
# Stage 1: per-modality converters
# ---------------------------------------------------------------------------

def convert_func(entry, sub, ses, xnat_dir, out_func_dir):
    bids_task = entry['bids_task']
    run       = entry['run']
    base      = f'sub-{sub}_ses-{ses}_task-{bids_task}_run-{run}'

    with tempfile.TemporaryDirectory(prefix='dcm2niix_func_') as tmp:
        merge_dicoms(xnat_dir, entry['xnat_scan_dirs'], tmp)
        run_dcm2niix(tmp, tmp)

        for fpath in glob.glob(os.path.join(tmp, '*.nii.gz')):
            fname = os.path.basename(fpath)
            part  = 'phase' if is_phase_file(fname) else 'mag'
            dest  = os.path.join(out_func_dir, f'{base}_part-{part}_bold.nii.gz')
            shutil.move(fpath, dest)
            print(f'    {part:5s}  →  {os.path.basename(dest)}')

        for fpath in glob.glob(os.path.join(tmp, '*.json')):
            fname = os.path.basename(fpath)
            part  = 'phase' if is_phase_file(fname) else 'mag'
            dest  = os.path.join(out_func_dir, f'{base}_part-{part}_bold.json')
            shutil.move(fpath, dest)
            inject_task_name(dest, bids_task)


def convert_physio(physio_folder, sub, ses, bids_task, run, xnat_dir, out_func_dir):
    """Copy physio files (.ima → .dcm) with BIDS-style naming into func/."""
    src = _dicom_dir(xnat_dir, physio_folder)
    if not os.path.isdir(src):
        print(f'    WARNING: physio DICOM dir not found: {src}')
        return

    base  = f'sub-{sub}_ses-{ses}_task-{bids_task}_run-{run}_physio'
    files = sorted(f for f in os.listdir(src) if os.path.isfile(os.path.join(src, f)))

    for i, fname in enumerate(files, 1):
        _, ext = os.path.splitext(fname)
        new_ext  = '.dcm' if ext.lower() == '.ima' else ext
        out_name = f'{base}_{i:03d}{new_ext}' if len(files) > 1 else f'{base}{new_ext}'
        shutil.copy2(os.path.join(src, fname), os.path.join(out_func_dir, out_name))

    print(f'    physio  →  {base}* ({len(files)} file(s))')


def convert_fmap(entry, sub, ses, xnat_dir, out_fmap_dir, out_func_dir):
    """Convert fieldmap; keep magnitude only in fmap/, physio (if any) in func/."""
    base = f'sub-{sub}_ses-{ses}_dir-{entry["dir"]}_epi'

    with tempfile.TemporaryDirectory(prefix='dcm2niix_fmap_') as tmp:
        merge_dicoms(xnat_dir, entry['xnat_scan_dirs'], tmp)
        run_dcm2niix(tmp, tmp)

        for fpath in glob.glob(os.path.join(tmp, '*.nii.gz')):
            if is_phase_file(os.path.basename(fpath)):
                continue   # discard phase for fieldmap
            dest = os.path.join(out_fmap_dir, f'{base}.nii.gz')
            shutil.move(fpath, dest)
            print(f'    mag    →  {os.path.basename(dest)}')

        for fpath in glob.glob(os.path.join(tmp, '*.json')):
            if is_phase_file(os.path.basename(fpath)):
                continue
            dest = os.path.join(out_fmap_dir, f'{base}.json')
            shutil.move(fpath, dest)

    if 'physio_dir' in entry:
        # Fieldmap physio kept alongside functional physio in func/
        fmap_base = f'sub-{sub}_ses-{ses}_dir-{entry["dir"]}_epi_physio'
        src = _dicom_dir(xnat_dir, entry['physio_dir'])
        if os.path.isdir(src):
            files = sorted(f for f in os.listdir(src) if os.path.isfile(os.path.join(src, f)))
            for i, fname in enumerate(files, 1):
                _, ext = os.path.splitext(fname)
                new_ext  = '.dcm' if ext.lower() == '.ima' else ext
                out_name = (f'{fmap_base}_{i:03d}{new_ext}' if len(files) > 1
                            else f'{fmap_base}{new_ext}')
                shutil.copy2(os.path.join(src, fname),
                             os.path.join(out_func_dir, out_name))
            print(f'    physio →  {fmap_base}* ({len(files)} file(s))')


def convert_anat(entry, sub, ses, xnat_dir, out_anat_dir):
    """Convert a single anatomical series (INV1, INV2, or UNI)."""
    bids_suffix = entry['bids_suffix']
    base        = f'sub-{sub}_ses-{ses}_{bids_suffix}'

    with tempfile.TemporaryDirectory(prefix='dcm2niix_anat_') as tmp:
        src = _dicom_dir(xnat_dir, entry['xnat_scan_dir'])
        if not os.path.isdir(src):
            print(f'    WARNING: DICOM dir not found: {src}')
            return
        shutil.copytree(src, os.path.join(tmp, 'dicoms'))
        run_dcm2niix(os.path.join(tmp, 'dicoms'), tmp)

        for fpath in glob.glob(os.path.join(tmp, '*.nii.gz')):
            dest = os.path.join(out_anat_dir, f'{base}.nii.gz')
            shutil.move(fpath, dest)
            print(f'    anat   →  {os.path.basename(dest)}')

        for fpath in glob.glob(os.path.join(tmp, '*.json')):
            dest = os.path.join(out_anat_dir, f'{base}.json')
            shutil.move(fpath, dest)

# ---------------------------------------------------------------------------
# Stage 1: orchestrator
# ---------------------------------------------------------------------------

def run_stage1(mapping, sub, ses, xnat_dir):
    sub_ses_dir  = os.path.join(SOURCEDATA_DIR, f'sub-{sub}', f'ses-{ses}')
    out_func_dir = os.path.join(sub_ses_dir, 'func')
    out_fmap_dir = os.path.join(sub_ses_dir, 'fmap')
    out_anat_dir = os.path.join(sub_ses_dir, 'anat')

    for d in [out_func_dir, out_fmap_dir, out_anat_dir]:
        os.makedirs(d, exist_ok=True)

    # ── Functional ────────────────────────────────────────────────────────
    print('\n[Stage 1] Functional runs:')
    for entry in mapping.get('func', []):
        tag = f'task-{entry["bids_task"]} run-{entry["run"]}'
        if not entry['include']:
            print(f'  {tag}  → skipped ({entry.get("reason", "failed")})')
            continue

        print(f'  {tag}')
        try:
            convert_func(entry, sub, ses, xnat_dir, out_func_dir)
        except RuntimeError as e:
            print(f'    ERROR during conversion: {e}')
            continue

        if 'physio_dir' in entry:
            convert_physio(
                entry['physio_dir'], sub, ses,
                entry['bids_task'], entry['run'],
                xnat_dir, out_func_dir,
            )

    # ── Fieldmaps ─────────────────────────────────────────────────────────
    print('\n[Stage 1] Fieldmaps:')
    for entry in mapping.get('fmap', []):
        print(f'  dir-{entry["dir"]}')
        try:
            convert_fmap(entry, sub, ses, xnat_dir, out_fmap_dir, out_func_dir)
        except RuntimeError as e:
            print(f'    ERROR during conversion: {e}')

    # ── Anatomicals ───────────────────────────────────────────────────────
    print('\n[Stage 1] Anatomicals:')
    for entry in mapping.get('anat', []):
        print(f'  {entry["bids_suffix"]}')
        try:
            convert_anat(entry, sub, ses, xnat_dir, out_anat_dir)
        except RuntimeError as e:
            print(f'    ERROR during conversion: {e}')

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_args()
    sub  = args.sub.zfill(2)
    ses  = args.ses.zfill(2)

    # Resolve XNAT session directory
    if args.xnat_dir:
        xnat_dir = os.path.abspath(args.xnat_dir)
        print(f'[Discovery] Using explicit XNAT dir: {xnat_dir}')
    else:
        print(f'[Discovery] Searching for sub-{sub} ses-{ses} in {XNAT_BASE_DIR}...')
        xnat_dir = find_xnat_dir(sub, ses)
        print(f'  Found: {os.path.basename(xnat_dir)}')

    sub_ses_dir = os.path.join(SOURCEDATA_DIR, f'sub-{sub}', f'ses-{ses}')
    json_path   = os.path.join(sub_ses_dir, 'scan_mapping.json')
    os.makedirs(sub_ses_dir, exist_ok=True)

    # ── Stage 0: build scan_mapping.json ──────────────────────────────────
    if os.path.exists(json_path) and not args.remap:
        print(f'\n[Stage 0] scan_mapping.json exists — loading '
              f'(use --remap to regenerate).')
        with open(json_path) as f:
            mapping = json.load(f)
    else:
        print('\n[Stage 0] Building scan_mapping.json...')

        excel_path = EXCEL_TEMPLATE.format(sub_int=int(sub))
        print(f'  Excel : {excel_path}')
        excel_runs = parse_excel(excel_path, ses)
        print(f'  Rows  : {len(excel_runs)} functional task entries found')

        print(f'  XNAT  : {os.path.join(xnat_dir, "scans")}')
        xnat_scans = inventory_xnat_scans(xnat_dir)
        print(f'  Scans : {len(xnat_scans)} folders inventoried')

        mapping = build_scan_mapping(sub, ses, xnat_dir, excel_runs, xnat_scans)

        with open(json_path, 'w') as f:
            json.dump(mapping, f, indent=2)
        print(f'  Written: {json_path}')

    if args.stage0_only:
        print('\n[Stage 0 only] Done.')
        return

    # ── Stage 1: DICOM → NIfTI ────────────────────────────────────────────
    print('\n[Stage 1] Converting DICOMs to NIfTI...')
    run_stage1(mapping, sub, ses, xnat_dir)

    print(f'\nDone. Outputs in: {sub_ses_dir}')


if __name__ == '__main__':
    main()
